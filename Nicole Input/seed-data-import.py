"""
Beauty Intelligence Platform — Seed Data Import Script
Reads the seed catalog Excel file and inserts into Supabase.

Prerequisites:
  pip install openpyxl supabase

Usage:
  1. Set environment variables:
     export SUPABASE_URL="https://your-project.supabase.co"
     export SUPABASE_SERVICE_KEY="your-service-role-key"
  
  2. Run the migration SQL first (supabase-migration.sql)
  
  3. Run this script:
     python seed-data-import.py --file beauty_intelligence_seed_catalog.xlsx

Note: Uses the service role key (not anon key) to bypass RLS for data import.
"""

import openpyxl
import json
import os
import sys
import argparse
from typing import Optional

# ═══════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")

# Zone mapping: subcategory → zone (for makeup products)
ZONE_MAP = {
    "Foundation": "Face", "Concealer": "Face", "Powder": "Face",
    "Primer": "Face", "Bronzer": "Face", "Contour": "Face",
    "Setting Spray": "Face",
    "Blush": "Cheek", "Highlighter": "Cheek",
    "Lip Color": "Lip", "Lip Gloss": "Lip", "Lip Liner": "Lip",
    "Tinted Lip Balm": "Lip",
    "Eyeshadow": "Eye", "Eyeliner": "Eye", "Mascara": "Eye",
    "Brow": "Eye",
}


# ═══════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════

def is_brand_row(value: str) -> bool:
    """Check if a row is a brand separator (── BRAND ──)"""
    return value.strip().startswith("──")

def is_data_row(value: str) -> bool:
    """Check if a row contains actual data"""
    v = value.strip()
    return bool(v) and not v.startswith("──")

def clean_value(value) -> Optional[str]:
    """Clean a cell value"""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None

def parse_array(value) -> list:
    """Parse a comma-separated string or existing array into a list"""
    if value is None:
        return []
    if isinstance(value, list):
        return value
    s = str(value).strip()
    if not s:
        return []
    return [item.strip() for item in s.split(",") if item.strip()]

def parse_price(value) -> Optional[float]:
    """Parse price value"""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        s = str(value).replace("$", "").strip()
        try:
            return float(s)
        except:
            return None

def get_headers(ws, row=2) -> dict:
    """Get column index mapping from header row"""
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=row, column=col).value
        if val:
            headers[str(val).strip()] = col
    return headers


# ═══════════════════════════════════════════════════════════
# SKINCARE IMPORT
# ═══════════════════════════════════════════════════════════

def import_skincare(ws) -> list:
    """Import skincare products from Skincare Catalog sheet"""
    headers = get_headers(ws)
    products = []
    
    for row in range(3, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        if hasattr(cell_a, 'value') is False:
            continue
        val = str(cell_a.value or "").strip()
        if not is_data_row(val):
            continue
        
        h = headers
        product = {
            "product_id": clean_value(ws.cell(row=row, column=h.get("Product ID", 1)).value),
            "brand": clean_value(ws.cell(row=row, column=h.get("Brand", 3)).value),
            "product_name": clean_value(ws.cell(row=row, column=h.get("Product Name", 4)).value),
            "category": "skincare",
            "subcategory": clean_value(ws.cell(row=row, column=h.get("Subcategory", 6)).value),
            "zone": None,
            "finish": None,
            "coverage": None,
            "size": None,
            "shade_count": 0,
            "primary_functions": [],
            "secondary_functions": [],
            "concentration_tier": clean_value(ws.cell(row=row, column=h.get("Concentration Tier", 11)).value),
            "treatment_or_support": clean_value(ws.cell(row=row, column=h.get("Treatment or Support", 13)).value),
            "core_from": None,
            "overlap_rule": None,
            "shade_relevance": None,
            "pairs_with": [],
            "formula_notes": None,
            "price": None,
            "upc": None,
            "status": "Complete",
        }
        
        # Parse functions
        pf1 = clean_value(ws.cell(row=row, column=h.get("Primary Function 1", 7)).value)
        pf2 = clean_value(ws.cell(row=row, column=h.get("Primary Function 2", 8)).value)
        product["primary_functions"] = [f for f in [pf1, pf2] if f]
        
        sf = clean_value(ws.cell(row=row, column=h.get("Secondary Functions", 9)).value)
        product["secondary_functions"] = parse_array(sf)
        
        # Formula notes = key ingredients + notes
        ingredients = clean_value(ws.cell(row=row, column=h.get("Key Active Ingredients", 10)).value)
        notes = clean_value(ws.cell(row=row, column=h.get("Notes", 18)).value)
        parts = [p for p in [ingredients, notes] if p]
        product["formula_notes"] = ". ".join(parts) if parts else None
        
        product["price"] = parse_price(ws.cell(row=row, column=h.get("Price", 14)).value)
        product["upc"] = clean_value(ws.cell(row=row, column=h.get("UPC", 15)).value)
        product["status"] = clean_value(ws.cell(row=row, column=h.get("Status", 17)).value) or "Complete"
        
        if product["product_id"]:
            products.append(product)
    
    return products


# ═══════════════════════════════════════════════════════════
# MAKEUP IMPORT
# ═══════════════════════════════════════════════════════════

def import_makeup(ws) -> list:
    """Import makeup products from Makeup Catalog sheet"""
    headers = get_headers(ws)
    products = []
    
    for row in range(3, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        val = str(cell_a.value or "").strip()
        if not is_data_row(val):
            continue
        
        h = headers
        subcategory = clean_value(ws.cell(row=row, column=h.get("Subcategory", 4)).value)
        
        product = {
            "product_id": clean_value(ws.cell(row=row, column=h.get("Product ID", 1)).value),
            "brand": clean_value(ws.cell(row=row, column=h.get("Brand", 2)).value),
            "product_name": clean_value(ws.cell(row=row, column=h.get("Product Name", 3)).value),
            "category": "makeup",
            "subcategory": subcategory,
            "zone": ZONE_MAP.get(subcategory),  # computed from subcategory
            "finish": clean_value(ws.cell(row=row, column=h.get("Finish", 5)).value),
            "coverage": clean_value(ws.cell(row=row, column=h.get("Coverage", 6)).value),
            "size": clean_value(ws.cell(row=row, column=h.get("Size", 7)).value),
            "shade_count": int(ws.cell(row=row, column=h.get("Shade Count", 8)).value or 0),
            "primary_functions": [],
            "secondary_functions": [],
            "concentration_tier": None,
            "treatment_or_support": None,
            "core_from": clean_value(ws.cell(row=row, column=h.get("Core From", 12)).value),
            "overlap_rule": clean_value(ws.cell(row=row, column=h.get("Overlap Rule", 13)).value),
            "shade_relevance": clean_value(ws.cell(row=row, column=h.get("Shade Relevance", 14)).value),
            "pairs_with": parse_array(ws.cell(row=row, column=h.get("Pairs With", 10)).value),
            "formula_notes": clean_value(ws.cell(row=row, column=h.get("Formula Notes", 9)).value),
            "price": parse_price(ws.cell(row=row, column=h.get("Price", 11)).value),
            "upc": None,
            "status": clean_value(ws.cell(row=row, column=h.get("Status", 15)).value) or "Complete",
        }
        
        if product["product_id"]:
            products.append(product)
    
    return products


# ═══════════════════════════════════════════════════════════
# SHADES IMPORT
# ═══════════════════════════════════════════════════════════

def import_shades(ws) -> list:
    """Import shade variants from Makeup Shades sheet"""
    headers = get_headers(ws)
    shades = []
    
    for row in range(3, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        val = str(cell_a.value or "").strip()
        if not is_data_row(val):
            continue
        
        h = headers
        shade = {
            "shade_id": clean_value(ws.cell(row=row, column=h.get("Shade ID", 1)).value),
            "product_id": clean_value(ws.cell(row=row, column=h.get("Parent ID", 2)).value),
            "shade_name": clean_value(ws.cell(row=row, column=h.get("Shade Name", 5)).value),
            "undertone": clean_value(ws.cell(row=row, column=h.get("Undertone", 6)).value),
            "color_family": clean_value(ws.cell(row=row, column=h.get("Color Family", 7)).value),
            "skin_depth_match": clean_value(ws.cell(row=row, column=h.get("Skin Depth Match", 8)).value),
            "flattering_range": clean_value(ws.cell(row=row, column=h.get("Flattering Range", 9)).value),
            "purpose": clean_value(ws.cell(row=row, column=h.get("Purpose", 10)).value),
            "finish": clean_value(ws.cell(row=row, column=h.get("Finish", 11)).value),
            "availability": clean_value(ws.cell(row=row, column=h.get("Availability", 12)).value) or "Permanent",
            "shade_description": clean_value(ws.cell(row=row, column=h.get("Shade Description", 13)).value),
            "primary_match": clean_value(ws.cell(row=row, column=h.get("Primary Match", 14)).value),
            "extended_match": clean_value(ws.cell(row=row, column=h.get("Extended Match", 15)).value),
            "status": clean_value(ws.cell(row=row, column=h.get("Status", 16)).value) or "Complete",
        }
        
        # Skip shade rows where Brand/Product Name columns exist but shade_id doesn't
        if shade["shade_id"]:
            shades.append(shade)
    
    return shades


# ═══════════════════════════════════════════════════════════
# BODY & HAIR IMPORT
# ═══════════════════════════════════════════════════════════

def import_body_hair(ws) -> list:
    """Import body and hair products from Body & Hair Catalog sheet"""
    headers = get_headers(ws)
    products = []
    
    for row in range(3, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        val = str(cell_a.value or "").strip()
        if not is_data_row(val):
            continue
        
        h = headers
        category_val = clean_value(ws.cell(row=row, column=h.get("Category", 5)).value)
        category = "body"
        if category_val and "hair" in category_val.lower():
            category = "hair"
        
        # Build formula notes from multiple fields
        treatment_styling = clean_value(ws.cell(row=row, column=h.get("Treatment or Styling", 8)).value)
        protein_moisture = clean_value(ws.cell(row=row, column=h.get("Protein or Moisture", 9)).value)
        heat = clean_value(ws.cell(row=row, column=h.get("Heat Protection", 10)).value)
        notes = clean_value(ws.cell(row=row, column=h.get("Notes", 13)).value)
        formula_parts = [p for p in [protein_moisture, heat, notes] if p]
        
        product = {
            "product_id": clean_value(ws.cell(row=row, column=h.get("Product ID", 1)).value),
            "brand": clean_value(ws.cell(row=row, column=h.get("Brand", 3)).value),
            "product_name": clean_value(ws.cell(row=row, column=h.get("Product Name", 4)).value),
            "category": category,
            "subcategory": clean_value(ws.cell(row=row, column=h.get("Subcategory", 6)).value),
            "zone": None,
            "finish": None,
            "coverage": None,
            "size": None,
            "shade_count": 0,
            "primary_functions": [clean_value(ws.cell(row=row, column=h.get("Primary Function", 7)).value)] if clean_value(ws.cell(row=row, column=h.get("Primary Function", 7)).value) else [],
            "secondary_functions": [],
            "concentration_tier": None,
            "treatment_or_support": "Treatment" if treatment_styling == "Treatment" else "Support",
            "core_from": None,
            "overlap_rule": None,
            "shade_relevance": None,
            "pairs_with": [],
            "formula_notes": ". ".join(formula_parts) if formula_parts else None,
            "price": parse_price(ws.cell(row=row, column=h.get("Price", 11)).value),
            "upc": None,
            "status": clean_value(ws.cell(row=row, column=h.get("Status", 12)).value) or "Complete",
        }
        
        if product["product_id"]:
            products.append(product)
    
    return products


# ═══════════════════════════════════════════════════════════
# CONFLICT RULES IMPORT
# ═══════════════════════════════════════════════════════════

def import_conflict_rules(ws) -> list:
    """Import conflict rules from Conflict Rules sheet"""
    headers = get_headers(ws)
    rules = []
    
    for row in range(3, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        val = str(cell_a.value or "").strip()
        if not val or val.startswith("──"):
            continue
        
        h = headers
        rule = {
            "rule_id": clean_value(ws.cell(row=row, column=h.get("Rule ID", 1)).value),
            "category_a": clean_value(ws.cell(row=row, column=h.get("Category A", 2)).value),
            "category_b": clean_value(ws.cell(row=row, column=h.get("Category B", 3)).value),
            "severity": clean_value(ws.cell(row=row, column=h.get("Severity", 4)).value),
            "condition": clean_value(ws.cell(row=row, column=h.get("Condition", 5)).value),
            "max_alignment": None,  # extract from engine behavior
            "explanation": clean_value(ws.cell(row=row, column=h.get("User-Facing Explanation", 7)).value),
            "resolutions": [],  # must be added manually — see Engine Logic Spec
        }
        
        # Extract max alignment from engine behavior text
        behavior = clean_value(ws.cell(row=row, column=h.get("Engine Behavior", 6)).value) or ""
        if "cannot be high" in behavior.lower():
            rule["max_alignment"] = "Moderate"
        elif "low" in behavior.lower():
            rule["max_alignment"] = "Low"
        
        if rule["rule_id"]:
            rules.append(rule)
    
    return rules


# ═══════════════════════════════════════════════════════════
# SUPABASE INSERT
# ═══════════════════════════════════════════════════════════

def insert_to_supabase(table: str, data: list, supabase_client):
    """Insert data into a Supabase table in batches"""
    BATCH_SIZE = 50
    inserted = 0
    errors = []
    
    for i in range(0, len(data), BATCH_SIZE):
        batch = data[i:i + BATCH_SIZE]
        try:
            result = supabase_client.table(table).upsert(batch).execute()
            inserted += len(batch)
        except Exception as e:
            errors.append(f"Batch {i//BATCH_SIZE}: {str(e)}")
    
    return inserted, errors


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Import seed catalog into Supabase")
    parser.add_argument("--file", required=True, help="Path to seed catalog Excel file")
    parser.add_argument("--dry-run", action="store_true", help="Parse data without inserting")
    parser.add_argument("--table", help="Import only a specific table: products, shades, conflict_rules")
    args = parser.parse_args()
    
    # Load workbook
    print(f"Loading {args.file}...")
    wb = openpyxl.load_workbook(args.file)
    print(f"Sheets found: {wb.sheetnames}")
    
    # Parse all data
    all_products = []
    all_shades = []
    all_rules = []
    
    if "Skincare Catalog" in wb.sheetnames:
        skincare = import_skincare(wb["Skincare Catalog"])
        all_products.extend(skincare)
        print(f"  Skincare: {len(skincare)} products")
    
    if "Makeup Catalog" in wb.sheetnames:
        makeup = import_makeup(wb["Makeup Catalog"])
        all_products.extend(makeup)
        print(f"  Makeup: {len(makeup)} products")
    
    if "Body & Hair Catalog" in wb.sheetnames:
        body_hair = import_body_hair(wb["Body & Hair Catalog"])
        all_products.extend(body_hair)
        print(f"  Body & Hair: {len(body_hair)} products")
    
    if "Makeup Shades" in wb.sheetnames:
        all_shades = import_shades(wb["Makeup Shades"])
        print(f"  Shades: {len(all_shades)} variants")
    
    if "Conflict Rules" in wb.sheetnames:
        all_rules = import_conflict_rules(wb["Conflict Rules"])
        print(f"  Conflict Rules: {len(all_rules)} rules")
    
    print(f"\nTotal: {len(all_products)} products, {len(all_shades)} shades, {len(all_rules)} rules")
    
    if args.dry_run:
        print("\n[DRY RUN] No data inserted. Sample output:")
        if all_products:
            print(f"\nFirst product:\n{json.dumps(all_products[0], indent=2)}")
        if all_shades:
            print(f"\nFirst shade:\n{json.dumps(all_shades[0], indent=2)}")
        if all_rules:
            print(f"\nFirst rule:\n{json.dumps(all_rules[0], indent=2)}")
        return
    
    # Connect to Supabase
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("\nERROR: Set SUPABASE_URL and SUPABASE_SERVICE_KEY environment variables.")
        sys.exit(1)
    
    from supabase import create_client
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    # Insert in order (products first, then shades which reference them)
    if not args.table or args.table == "products":
        print(f"\nInserting {len(all_products)} products...")
        inserted, errors = insert_to_supabase("products", all_products, supabase)
        print(f"  Inserted: {inserted}")
        if errors:
            print(f"  Errors: {errors}")
    
    if not args.table or args.table == "shades":
        print(f"\nInserting {len(all_shades)} shades...")
        inserted, errors = insert_to_supabase("shades", all_shades, supabase)
        print(f"  Inserted: {inserted}")
        if errors:
            print(f"  Errors: {errors}")
    
    if not args.table or args.table == "conflict_rules":
        print(f"\nInserting {len(all_rules)} conflict rules...")
        inserted, errors = insert_to_supabase("conflict_rules", all_rules, supabase)
        print(f"  Inserted: {inserted}")
        if errors:
            print(f"  Errors: {errors}")
    
    print("\nDone.")


if __name__ == "__main__":
    main()
